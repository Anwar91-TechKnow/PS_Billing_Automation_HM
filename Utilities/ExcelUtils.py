import openpyxl
from openpyxl import load_workbook

workbook = openpyxl.load_workbook(
    "C://Users//anwarshaikh//PycharmProjects//PS_Billing_Automation//TestData//Engagement_Summary_Sheet.xlsx")
sheet = workbook['chatinfo']

total_rows = sheet.max_row
total_cols = sheet.max_column


for r in range(1,total_rows+1):
    for c in range(1,total_cols+1):
        print(sheet.cell(r,c).value)







print(sheet.cell(1,1).value)
# # print(total_cols)
# print(total_rows)